import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
import pandas as pd
import pdfplumber
import re
from openpyxl.utils import get_column_letter
import os
import numpy as np
import easyocr
import warnings
import contextlib
import sys

warnings.filterwarnings("ignore")

# Redirecionar print para o log do Tkinter
class TextRedirector:
    def __init__(self, widget):
        self.widget = widget

    def write(self, string):
        self.widget.configure(state='normal')
        self.widget.insert(tk.END, string)
        self.widget.see(tk.END)
        self.widget.configure(state='disabled')

    def flush(self):
        pass

@contextlib.contextmanager
def suppress_stdout():
    with open(os.devnull, "w") as devnull:
        old_stdout = sys.stdout
        sys.stdout = devnull
        try:
            yield
        finally:
            sys.stdout = old_stdout

# DataFrame global
df = None

# Função para log
def log(msg):
    log_text.configure(state='normal')
    log_text.insert(tk.END, msg + "\n")
    log_text.see(tk.END)
    log_text.configure(state='disabled')

# Função para processar área de interesse
def processar_area(area_interesse):
    global df
    area_interesse = re.sub(r"\s+", " ", area_interesse).strip()
    ocorrencias = [o.strip() for o in area_interesse.split(';') if o.strip()]

    resultado = []
    for occ in ocorrencias:
        ids = re.findall(r'\b\d{15}\b', occ)
        if ids:
            primeiro_id = ids[0]
            descricao = occ.split(primeiro_id)[0].rstrip(': ').strip()
            medicamentos = []
            partes = re.split(r'\b\d{15}\b', occ)
            for p in partes[1:]:
                p = p.strip(" ,")
                if p:
                    medicamentos.append(p)
        else:
            descricao = occ
            medicamentos = []
        resultado.append({
            "Ocorrência": descricao,
            "Autorização": ids,
            "Medicamentos": medicamentos
        })

    linhas = []
    for occ in resultado:
        descricao = occ["Ocorrência"]
        ids = occ["Autorização"]
        medicamentos = occ["Medicamentos"]
        while len(medicamentos) < len(ids):
            medicamentos.append('')
        for idx, id_aut in enumerate(ids):
            med = medicamentos[idx].strip()
            texto_final = f"{descricao} {med}".strip()
            linhas.append({
                "Autorização": id_aut,
                "Ocorrência": texto_final
            })

    df = pd.DataFrame(linhas)
    log("Área de interesse processada com sucesso!")
    messagebox.showinfo("Sucesso", "Área de interesse processada com sucesso!")

# Função para processar PDF
def processar_pdf():
    global df
    log_text.configure(state='normal')
    log_text.delete(1.0, tk.END)
    log_text.configure(state='disabled')

    arquivo = filedialog.askopenfilename(
        title="Selecione o arquivo PDF",
        filetypes=[("PDF", "*.pdf")]
    )

    if not arquivo:
        return

    try:
        log("Iniciando extração de texto do PDF...")

        # ===============================
        # 1️⃣ Tentar extrair texto com pdfplumber
        # ===============================
        texto = ""
        with pdfplumber.open(arquivo) as pdf:
            for i, pagina in enumerate(pdf.pages):
                pagina_texto = pagina.extract_text()
                if pagina_texto:
                    texto += pagina_texto + "\n"
                log(f"Página {i+1} processada com pdfplumber.")

        texto = re.sub(r"\s+", " ", texto).strip()

        # ===============================
        # 2️⃣ Se pdfplumber falhar, usar OCR com EasyOCR
        # ===============================
        if not texto:
            log("Texto não encontrado. Iniciando OCR com EasyOCR...")
            with suppress_stdout():
                reader = easyocr.Reader(['pt'])

            texto_paginas = []
            with pdfplumber.open(arquivo) as pdf:
                total_paginas = len(pdf.pages)
                progress["maximum"] = total_paginas
                progress["value"] = 0

                for i, page in enumerate(pdf.pages):
                    pil_image = page.to_image(resolution=300).original
                    image_np = np.array(pil_image)
                    resultado = reader.readtext(image_np, detail=0)
                    texto_paginas.append("\n".join(resultado))
                    log(f"Página {i+1} processada com OCR.")
                    progress["value"] = i + 1
                    root.update_idletasks()

            texto = "\n\n".join(texto_paginas)
            progress["value"] = 0

        if not texto:
            messagebox.showwarning("Aviso", "Não foi possível extrair texto do PDF.")
            return

        log("Extração de texto concluída. Processando ocorrências...")

        matches = list(re.finditer(r"\d{15}", texto))
        if not matches:
            messagebox.showwarning("Aviso", "Nenhum número de 15 dígitos encontrado no PDF.")
            return

        start_idx = matches[0].start()
        end_idx = matches[-1].end()

        # Localiza todos os ":" antes do primeiro número de 15 dígitos
        all_colons = [m.start() for m in re.finditer(r":", texto)]
        colons_before = [c for c in all_colons if c < start_idx]

        # Início da área: segundo ":" antes do primeiro número, ou o primeiro, ou 0
        cut_start = colons_before[-2]+1 if len(colons_before) >= 2 else (colons_before[0]+1 if colons_before else 0)
        cut_end = matches[-1].end()
        area_interesse = texto[cut_start:cut_end].strip()

        processar_area(area_interesse)

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro ao processar o PDF:\n{e}")

# Função para abrir janela de input de texto
def abrir_input_texto():
    input_win = tk.Toplevel(root)
    input_win.title("Colar Texto")
    input_win.geometry("500x400")
    input_win.resizable(False, False)

    tk.Label(input_win, text="Cole aqui o texto:").pack(pady=5)

    text_area = scrolledtext.ScrolledText(input_win, width=60, height=15)
    text_area.pack(pady=5)

    def processar_colado():
        conteudo = text_area.get("1.0", tk.END).strip()
        if not conteudo:
            messagebox.showwarning("Atenção", "Nenhum texto encontrado.")
            return
        input_win.destroy()
        log_text.configure(state='normal')
        log_text.delete(1.0, tk.END)
        log_text.configure(state='disabled')
        processar_area(conteudo)

    tk.Button(input_win, text="Processar", command=processar_colado, width=20, height=2).pack(pady=10)

# Função para salvar DataFrame em Excel
def salvar_excel():
    global df
    if df is None or df.empty:
        messagebox.showwarning("Atenção", "Nenhum dado para salvar. Processar PDF ou colar texto primeiro.")
        return

    caminho_saida = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")],
        title="Salvar como Excel"
    )

    if not caminho_saida:
        return

    try:
        with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Ocorrências")
            ws = writer.sheets["Ocorrências"]
            for col_idx, col in enumerate(df.columns, 1):
                col_letter = get_column_letter(col_idx)
                max_length = max(df[col].astype(str).map(len).max(), len(col))
                ws.column_dimensions[col_letter].width = max_length + 2
        messagebox.showinfo("Sucesso", f"Arquivo salvo em:\n{caminho_saida}")
    except Exception as e:
        messagebox.showerror("Erro", f"Não foi possível salvar o Excel:\n{e}")

# Função para consolidar ocorrências
def consolidar_ocorrencias():
    global df

    if df is None or df.empty:
        messagebox.showwarning("Atenção", "Nenhum dado carregado para consolidar.")
        return

    try:
        # Normalizar nomes das colunas
        df.columns = df.columns.str.strip().str.lower()
        if 'ocorrência' not in df.columns and 'ocorrências' not in df.columns:
            messagebox.showerror("Erro", f"A coluna 'Ocorrência(s)' não foi encontrada.")
            return

        coluna = 'ocorrência' if 'ocorrência' in df.columns else 'ocorrências'
        df[coluna] = df[coluna].astype(str)

        # Explodir múltiplas ocorrências
        df_explodido = df[coluna].str.split(';').explode().str.strip()
        df_explodido = df_explodido[df_explodido != '']
        df_explodido = df_explodido[df_explodido.str.lower() != 'total']

        # Consolidar padrões específicos
        df_explodido = df_explodido.replace({
            r'^Intercambialidade incorreta do medicamento.*': 'Intercambialidade incorreta do medicamento',
            r'^Ausência da posologia do medicamento.*': 'Ausência da posologia do medicamento'
        }, regex=True)

        # Contar ocorrências
        df_consolidado = df_explodido.value_counts().reset_index()
        df_consolidado.columns = ['OCORRÊNCIAS', 'QTD']
        df_consolidado = df_consolidado.sort_values('OCORRÊNCIAS')

        # Adicionar total geral
        total = df_consolidado['QTD'].sum()
        df_consolidado.loc[len(df_consolidado)] = ['Total Geral', total]

        # Escolher local para salvar
        caminho_saida = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Arquivo Excel", "*.xlsx")],
            title="Salvar arquivo consolidado como"
        )

        if not caminho_saida:
            return

        # Salvar Excel
        with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
            df_consolidado.to_excel(writer, sheet_name="Ocorrências", index=False)
            ws = writer.sheets["Ocorrências"]
            for col_idx, col in enumerate(df_consolidado.columns, 1):
                col_letter = get_column_letter(col_idx)
                max_length = max(df_consolidado[col].astype(str).map(len).max(), len(col))
                ws.column_dimensions[col_letter].width = max_length + 2

        messagebox.showinfo("Sucesso", f"Arquivo consolidado salvo em:\n{caminho_saida}")

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro durante a consolidação:\n{e}")

# Interface Tkinter
root = tk.Tk()
root.title("Processador de PDF / Texto")
root.geometry("550x550")
root.resizable(False, False)

botao_style = {"width": 40, "height": 2, "padx": 5, "pady": 5}

tk.Button(root, text="Selecionar PDF e Processar", command=processar_pdf, **botao_style).pack(pady=5)
tk.Button(root, text="Colar Texto e Processar", command=abrir_input_texto, **botao_style).pack(pady=5)
tk.Button(root, text="Salvar em Excel", command=salvar_excel, **botao_style).pack(pady=5)
tk.Button(root, text="Consolidar Ocorrências", command=consolidar_ocorrencias, **botao_style).pack(pady=5)

# Barra de progresso
progress = ttk.Progressbar(root, orient="horizontal", length=500, mode="determinate")
progress.pack(pady=5)

# Área de log
log_text = scrolledtext.ScrolledText(root, width=65, height=15, state='disabled')
log_text.pack(pady=10)

root.mainloop()
