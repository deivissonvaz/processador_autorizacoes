import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import subprocess
import pandas as pd
import re
import numpy as np
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
from openpyxl.styles import Alignment, Font

# Dataframe global
df = None  # Variável global para armazenar o DataFrame

# Funções essenciais

def detectar_cabecalho(df_raw, min_validos=2):
    for i, row in df_raw.iterrows():
        if row.notna().sum() >= min_validos:
            df_raw.columns = row
            return df_raw.iloc[i+1:].reset_index(drop=True)
    raise ValueError("Cabeçalho não encontrado com base nos dados.")

def selecionar_planilha():
    global df

    root = tk.Tk()
    root.withdraw()

    caminho_arquivo = filedialog.askopenfilename(
        title="Selecione a planilha Excel",
        filetypes=[("Planilhas Excel", "*.xlsx")]
    )

    if not caminho_arquivo:
        return

    try:
        xls = pd.ExcelFile(caminho_arquivo)
        abas = xls.sheet_names
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao abrir o arquivo:\n{e}")
        return

    # Subjanela com Combobox
    subjanela = tk.Toplevel()
    subjanela.title("Selecionar aba")
    subjanela.geometry("320x140")
    subjanela.resizable(False, False)
    subjanela.grab_set()

    tk.Label(subjanela, text="Escolha a aba com os dados:").pack(pady=(15, 5))

    aba_escolhida = tk.StringVar()
    combo = ttk.Combobox(subjanela, textvariable=aba_escolhida, values=abas, state="readonly", width=30)
    combo.current(0)
    combo.pack(pady=5)

    def confirmar():
        global df
        try:
            aba = aba_escolhida.get()
            df_raw = pd.read_excel(caminho_arquivo, sheet_name=aba, header=None)
            df_limp = detectar_cabecalho(df_raw)

            colunas_esperadas = ["Datas", "Autorizações", "Ocorrências", "Valor Pago pelo MS"]
            df_limp.columns = colunas_esperadas[:len(df_limp.columns)]

            # Conversões de tipo
            df_limp["Datas"] = pd.to_datetime(df_limp["Datas"], errors='coerce').dt.date
            df_limp["Autorizações"] = pd.to_numeric(df_limp["Autorizações"], errors='coerce').fillna(0).astype(int)
            df_limp["Ocorrências"] = df_limp["Ocorrências"].astype(str)
            df_limp["Valor Pago pelo MS"] = pd.to_numeric(df_limp["Valor Pago pelo MS"], errors='coerce').round(2)

            # 🔧 Tratamento silencioso dos dados
            df_limp["Ocorrências"] = df_limp["Ocorrências"].str.replace(r';\s*;', ';', regex=True)
            df_limp = df_limp.dropna().reset_index(drop=True)

            df = df_limp.copy()
            messagebox.showinfo("Sucesso", f"Aba '{aba}' carregada e tratada com sucesso.")
            subjanela.destroy()
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao processar os dados:\n{e}")

    ttk.Button(subjanela, text="OK", command=confirmar).pack(pady=10)
    subjanela.mainloop()

def consolidar_pagamentos():
    """
    Consolida pagamentos do DataFrame global `df`, agrupando por mês/ano.
    Exporta os dados para um Excel formatado com moeda e colunas ajustadas.
    """
    global df

    if df is None:
        messagebox.showwarning("Atenção", "Nenhum dado carregado. Use 'Selecionar Planilha' primeiro.")
        return

    try:
        # Padronizar colunas
        df.columns = df.columns.str.strip().str.lower()

        # Conversão de datas
        df['datas'] = pd.to_datetime(df['datas'], errors='coerce')

        # Criar coluna de competência
        df['competência mês/ano numérica'] = df['datas'].dt.strftime('%m/%Y')
        

        # Agrupar por competência e somar valores
        df_consolidado = (
            df.groupby('competência mês/ano')['valor pago pelo ms']
            .sum()
            .reset_index(name='valor')
        )

        # Ordenar por data real
        df_consolidado['data referência'] = pd.to_datetime(df_consolidado['competência mês/ano'], format='%m/%Y')
        df_consolidado = df_consolidado.sort_values('data referência').drop(columns='data referência')

        # Adicionar linha de total
        total = df_consolidado['valor'].sum()
        df_consolidado.loc[len(df_consolidado)] = ['Total Geral', total]

        # Escolher local para salvar
        caminho_saida = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Arquivo Excel", "*.xlsx")],
            title="Salvar arquivo consolidado como"
        )

        if not caminho_saida:
            return  # Usuário cancelou

        # Salvar Excel
        with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
            df_consolidado.to_excel(writer, sheet_name="Consolidado", index=False)

            # Formatar colunas
            wb = writer.book
            ws = writer.sheets["Consolidado"]

            for col_idx, col in enumerate(df_consolidado.columns, 1):
                col_letter = get_column_letter(col_idx)

                # Ajustar largura da coluna
                max_length = max(
                    df_consolidado[col].astype(str).map(len).max(),
                    len(col)
                )
                ws.column_dimensions[col_letter].width = max_length + 2

                # Formatar como moeda se for a coluna 'valor'
                if col.lower() == "valor":
                    for row in range(2, ws.max_row + 1):  # pular cabeçalho
                        ws[f"{col_letter}{row}"].number_format = 'R$ #,##0.00'

        messagebox.showinfo("Sucesso", f"Arquivo salvo com sucesso em:\n{caminho_saida}")

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro durante a consolidação:\n{e}")

def consolidar_ocorrencias():
    """
    Consolida o número de ocorrências por mês/ano a partir do DataFrame global `df`.
    Exporta os dados consolidados para um arquivo Excel com colunas ajustadas.
    """
    global df

    if df is None:
        messagebox.showwarning("Atenção", "Nenhum dado carregado. Use 'Selecionar Planilha' primeiro.")
        return

    try:
        # Normalizar nomes das colunas
        df.columns = df.columns.str.strip().str.lower()

        if 'ocorrências' not in df.columns:
            messagebox.showerror("Erro", f"A coluna Ocorrências não foi encontrada no arquivo.")
            return

        # Garantir que os dados são string
        df['ocorrências'] = df['ocorrências'].astype(str)

        # Separar múltiplas ocorrências
        df_explodido = df['ocorrências'].str.split(';').explode().str.strip()

        # Remover entradas vazias e "Total"
        df_explodido = df_explodido[df_explodido.str.lower() != 'total']
        df_explodido = df_explodido[df_explodido != '']

        # Consolidar padrões específicos
        df_explodido = df_explodido.replace({
            r'^Intercambialidade incorreta do medicamento.*': 'Intercambialidade incorreta do medicamento',
            r'^Ausência da posologia do medicamento.*': 'Ausência da posologia do medicamento'
        }, regex=True)

        # Contar ocorrências
        df_consolidado = df_explodido.value_counts().reset_index()
        df_consolidado.columns = ['OCORRÊNCIAS', 'QTD']
        df_consolidado = df_consolidado.sort_values('OCORRÊNCIAS')

        # Adicionar total geral
        total = df_consolidado['QTD'].sum()
        df_consolidado.loc[len(df_consolidado)] = ['Total Geral', total]

        # Escolher local para salvar
        caminho_saida = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Arquivo Excel", "*.xlsx")],
            title="Salvar arquivo consolidado como"
        )

        if not caminho_saida:
            return  # Usuário cancelou

        # Salvar Excel
        with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
            df_consolidado.to_excel(writer, sheet_name="Ocorrências", index=False)

            # Formatar colunas
            wb = writer.book
            ws = writer.sheets["Ocorrências"]

            for col_idx, col in enumerate(df_consolidado.columns, 1):
                col_letter = get_column_letter(col_idx)

                # Ajustar largura da coluna
                max_length = max(
                    df_consolidado[col].astype(str).map(len).max(),
                    len(col)
                )
                ws.column_dimensions[col_letter].width = max_length + 2

        messagebox.showinfo("Sucesso", f"Arquivo salvo com sucesso em:\n{caminho_saida}")

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro durante a consolidação:\n{e}")

# Criando a janela principal
root = tk.Tk()
root.title("Painel de Controle")
root.geometry("300x250")
root.resizable(False, False)

# Estilo de botões
botao_style = {
    "width": 30,
    "height": 2,
    "padx": 5,
    "pady": 5
}

# Botões
btn_selecionar = tk.Button(root, text="Selecionar Planilha", command=selecionar_planilha, **botao_style)
btn_pagamentos = tk.Button(root, text="Consolidar Pagamentos", command=consolidar_pagamentos, **botao_style)
btn_ocorrencias = tk.Button(root, text="Consolidar Ocorrências", command=consolidar_ocorrencias, **botao_style)

# Layout
btn_selecionar.pack(pady=5)
btn_pagamentos.pack(pady=5)
btn_ocorrencias.pack(pady=5)

# Inicia a aplicação
root.mainloop()
